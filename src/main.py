import json
import logging
import os
from datetime import datetime
from enum import Enum
from pathlib import Path

from dotenv import load_dotenv
from openai import OpenAI
from pydantic import BaseModel, Field
from tenacity import retry, stop_after_attempt, wait_exponential

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configuration
INPUT_PATH = os.getenv("INPUT_PATH", "initial_data")
OUTPUT_PATH = os.getenv("OUTPUT_PATH", "output")
MODEL_NAME = os.getenv("MODEL_NAME", "gpt-4o")
EXPERT_TEMPERATURE = float(os.getenv("EXPERT_TEMPERATURE", "0.3"))
RETRY_ATTEMPTS = int(os.getenv("RETRY_ATTEMPTS", "3"))

APP_URL_ROUTES_FILE_NAME = os.getenv("APP_URL_ROUTES_FILE_NAME", "app_url_routes.json")
APP_DESCRIPTION_FILE_NAME = os.getenv(
    "APP_DESCRIPTION_FILE_NAME", "app_description.txt"
)


class TestPerspective(Enum):
    FUNCTIONAL = "functional"
    SECURITY = "security"
    ACCESSIBILITY = "accessibility"
    PERFORMANCE = "performance"
    USABILITY = "usability"


class PerspectiveConfig(BaseModel):
    name: TestPerspective
    temperature: float
    system_prompt: str


class ElementContext(BaseModel):
    element_id: str
    element_type: str
    element_data: dict
    surrounding_html: str  # HTML context around the element


TEST_CASES_DEFINITION_PROMPT = """
        Generate test cases with:
        - title: Clear description of the test
        - steps: Detailed and granular reproduction steps
        - result: Expected behaviors and validations
        - priority: High/Medium/Low
        - category: {perspective}
        - estimated_time: Expected execution time (quick/medium/long)
    
    Example of one test case for functional category:
      {{
        "title": "Verify that Event can be created",
        "steps": [
            "Navigate to Event page",
            "Click "Add event" page ",
            "Upload any valid WEBP, JPG or PNG image with up to 300 KB size at "Event image" section",
            "Fill in "Event Name" field with valid data, e.g. 'eventName'",
            "Pick up any valid date range at "Duration" date-picker",
            "Fill in "Event URL" field with valid url icnluding "https://" prefix",
            "Select any checkbox at "Appearance type" section",
            "Fill in "Short Description" field with text with min 100 characters and max 500",
            "Click "Create Event" button",
            "Check that all entered data during creation is applied"
        ],
        "result": [
            ""Events & Resources" page should be opened",
            "Created event should be displayed in the list",
            "All entered during creation data should be saved"
        ],
        "priority": "High",
        "category": "functional",
        "estimated_time": "medium"
    }}
"""

TEST_PERSPECTIVES = [
    PerspectiveConfig(
        name=TestPerspective.FUNCTIONAL,
        temperature=0.7,
        system_prompt="You are a QA expert focused on functional testing. Generate test cases that verify core functionality, user flows, and business logic.",
    ),
    # PerspectiveConfig(
    #     name=TestPerspective.SECURITY,
    #     temperature=0.8,
    #     system_prompt="You are a security testing expert. Generate test cases that identify potential security vulnerabilities, validate input sanitization, and verify access controls."
    # ),
    # PerspectiveConfig(
    #     name=TestPerspective.ACCESSIBILITY,
    #     temperature=0.6,
    #     system_prompt="You are an accessibility testing expert. Generate test cases that verify WCAG compliance, screen reader compatibility, and keyboard navigation."
    # ),
    # PerspectiveConfig(
    #     name=TestPerspective.PERFORMANCE,
    #     temperature=0.7,
    #     system_prompt="You are a performance testing expert. Generate test cases that verify load times, resource usage, and responsiveness under different conditions."
    # ),
    # PerspectiveConfig(
    #     name=TestPerspective.USABILITY,
    #     temperature=0.8,
    #     system_prompt="You are a usability testing expert. Generate test cases that verify user experience, interface consistency, and error handling from a user's perspective."
    # )
]


class TestMetadata(BaseModel):
    timestamp: str = Field(default_factory=lambda: datetime.now().isoformat())
    model_version: str
    temperature: float
    input_tokens: int = 0
    output_tokens: int = 0
    total_tokens: int = 0


class GeneratedTestCase(BaseModel):
    title: str
    steps: list[str]
    result: list[str]
    priority: str
    category: str
    estimated_time: str


class TestGenerationResult(BaseModel):
    test_cases: list[GeneratedTestCase]


class ExpertValidatedTests(BaseModel):
    test_cases: list[GeneratedTestCase]


class TestGenerationResultWithMetadata(BaseModel):
    result: TestGenerationResult
    metadata: TestMetadata


class ExpertValidatedTestsWithMetadata(BaseModel):
    result: ExpertValidatedTests
    validation_metadata: TestMetadata  # metadata from expert validation


class TestCaseGenerator:
    def __init__(self, initial_data_path: str):
        self.base_path = Path(initial_data_path)
        self.routes = self._load_json(APP_URL_ROUTES_FILE_NAME)
        self.app_description = self._load_text(APP_DESCRIPTION_FILE_NAME)
        self._client = OpenAI()

    def _load_json(self, filename: str) -> dict:
        """Load and parse JSON file"""
        try:
            with open(self.base_path / filename, "r") as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Error loading {filename}: {e}")
            return {}

    def _load_text(self, filename: str) -> str:
        """Load text file content"""
        try:
            with open(self.base_path / filename, "r") as f:
                return f.read()
        except Exception as e:
            logger.error(f"Error loading {filename}: {e}")
            return ""

    def _load_page_data(self, page_dir: str) -> dict:
        """Load all data for a specific page"""
        dir_path = self.base_path / page_dir
        data = {"html": "", "elements": ""}

        # Load HTML content
        html_files = list(dir_path.glob("*.html"))
        if len(html_files) == 1:
            with open(html_files[0], "r") as f:
                data["html"] = f.read()
        else:
            logger.warning(f"Found {len(html_files)} HTML files for {page_dir}")

        # Load JSON content
        json_files = list(dir_path.glob("*.json"))
        if len(json_files) == 1:
            data["elements"] = self._load_json(
                str(json_files[0].relative_to(self.base_path))
            )
        else:
            logger.warning(f"Found {len(json_files)} JSON files for {page_dir}")

        # Load page description
        desc_file = dir_path / "page_description.txt"
        if desc_file.exists():
            data["description"] = self._load_text(
                str(desc_file.relative_to(self.base_path))
            )

        return data

    def _extract_page_paths(self, routes: dict) -> list[str]:
        """Extract all possible page paths from routes configuration"""
        paths = []

        def process_route(route: dict, parent_path: str = ""):
            current_path = route["path"].strip("/")

            # Calculate full path
            if current_path:  # If it's not root or empty
                full_path = (
                    f"{parent_path}_{current_path}" if parent_path else current_path
                )
                paths.append(full_path)
            else:  # For root path "/"
                full_path = parent_path  # Keep parent path without adding anything

            # Process children regardless of current path
            if "children" in route:
                for child in route["children"]:
                    # For child routes, use the full_path as parent
                    process_route(child, full_path)

        # Process both private and public routes
        for route_type in ["privateRoutes", "publicRoutes"]:
            if route_type in routes:
                for route in routes[route_type]:
                    process_route(route)

        return paths

    @retry(
        stop=stop_after_attempt(RETRY_ATTEMPTS),
        wait=wait_exponential(multiplier=1, min=4, max=60),
    )
    def _generate_test_cases_with_retry(
        self, prompt: str, temperature: float
    ) -> TestGenerationResultWithMetadata:
        """Generate test cases using OpenAI API with retry logic"""
        try:
            response = self._client.beta.chat.completions.parse(
                model=MODEL_NAME,
                messages=[
                    {
                        "role": "system",
                        "content": "You are a QA automation expert. Generate detailed test cases in JSON format.",
                    },
                    {"role": "user", "content": prompt},
                ],
                response_format=TestGenerationResult,
                temperature=temperature,
            )

            logger.info("Usage info")
            openai_api_calculate_cost(response.usage)

            # Create metadata
            metadata = TestMetadata(
                model_version=MODEL_NAME,
                temperature=temperature,
                input_tokens=response.usage.prompt_tokens,
                output_tokens=response.usage.completion_tokens,
                total_tokens=response.usage.total_tokens,
            )

            # Parse the response
            parsed_response = json.loads(response.choices[0].message.content)
            # Extract the test_cases list from the response
            test_cases = parsed_response.get("test_cases", [])
            result = TestGenerationResult(test_cases=test_cases)

            return TestGenerationResultWithMetadata(result=result, metadata=metadata)

        except Exception as e:
            logger.error(f"Error generating test cases: {e}")
            raise

    def _validate_with_expert(
        self, test_results: list[TestGenerationResultWithMetadata], page_data: dict
    ) -> ExpertValidatedTestsWithMetadata:
        """Have an expert LLM validate and curate the test cases"""
        # Prepare the prompt for the expert
        expert_prompt = self._create_expert_prompt(test_results, page_data)

        try:
            response = self._client.beta.chat.completions.parse(
                model=MODEL_NAME,
                messages=[
                    {
                        "role": "system",
                        "content": """You are a senior QA expert. Your task is to:
                    1. Review and validate test cases
                    2. Remove duplicates and unnecessary tests
                    3. Ensure proper categorization and priorities
                    4. Verify estimated execution times
                    5. Enhance test steps and expected results if needed""",
                    },
                    {"role": "user", "content": expert_prompt},
                ],
                response_format=ExpertValidatedTests,
                temperature=EXPERT_TEMPERATURE,
            )

            logger.info("Usage info")
            openai_api_calculate_cost(response.usage)

            # Create validation metadata
            validation_metadata = TestMetadata(
                model_version=MODEL_NAME,
                temperature=EXPERT_TEMPERATURE,
                input_tokens=response.usage.prompt_tokens,
                output_tokens=response.usage.completion_tokens,
                total_tokens=response.usage.total_tokens,
            )

            parsed_response = json.loads(response.choices[0].message.content)
            validated_test_cases = parsed_response.get("test_cases", [])
            return ExpertValidatedTestsWithMetadata(
                result=ExpertValidatedTests(test_cases=validated_test_cases),
                validation_metadata=validation_metadata,
            )

        except Exception as e:
            logger.error(f"Error in expert validation: {e}")
            raise

    def _create_expert_prompt(
        self, test_results: list[TestGenerationResultWithMetadata], page_data: dict
    ) -> str:
        """Create prompt for expert validation"""
        # Group test cases by perspective for better organization
        test_cases_by_perspective = {}
        for result in test_results:
            for test in result.result.test_cases:
                perspective = test.category
                if perspective not in test_cases_by_perspective:
                    test_cases_by_perspective[perspective] = []
                test_cases_by_perspective[perspective].append(test)

        return f"""
        As a senior QA expert, review and validate test cases for a web page with the following context:

        PAGE CONTEXT:
        -------------
        Description: {page_data.get('description', 'N/A')}
        
        Form Elements: {json.dumps(page_data.get('elements', {}), indent=2)}
        
        HTML Content: {page_data.get('html', 'N/A')}

        GENERATED TEST CASES BY PERSPECTIVE:
        ----------------------------------
        {json.dumps({perspective: [test.model_dump() for test in tests] 
                    for perspective, tests in test_cases_by_perspective.items()}, indent=2)}

        YOUR TASK:
        ----------
        1. Review all test cases in the context of the actual page content and functionality
        
        2. For each testing perspective (functional, security, accessibility, etc.):
           - Ensure comprehensive coverage of that perspective's concerns
           - Remove redundant or overlapping tests within the perspective
           - Verify test cases align with best practices for that type of testing
        
        3. Analyze and optimize test case relationships:
           - Identify dependencies between tests
           - Ensure logical test execution order
           - Combine related tests when appropriate
           - Split complex tests into smaller, focused ones
        
        4. For each test case:
           - Improve steps clarity based on actual page elements
           - Make expected results more specific and verifiable
           - Verify priority matches the feature's importance
           - Validate estimated execution times
           - Ensure proper categorization
        
        5. Add critical test cases that are missing, considering:
           - Edge cases and error scenarios
           - Cross-perspective interactions
           - User flow completeness
           - Common failure patterns

        6. For element-specific test cases:
           - Verify they consider the element's context
           - Ensure coverage of element interactions
           - Check for proper state handling

        Provide your curated and enhanced test cases in the specified JSON format.
        Focus on creating a comprehensive yet practical test suite that maximizes coverage while minimizing redundancy.
        """

    def _get_perspective_specific_guidance(self, perspective: PerspectiveConfig) -> str:
        """Get specific testing guidance based on the perspective"""
        guidance = {
            TestPerspective.FUNCTIONAL: """
                - Test all core business functions
                - Verify form submissions and data processing
                - Check navigation flows
                - Validate business rules and logic
                - Test input validation and error handling
            """,
            TestPerspective.SECURITY: """
                - Test input sanitization
                - Check for XSS vulnerabilities
                - Verify authentication flows
                - Test authorization and access controls
                - Check for sensitive data exposure
            """,
            TestPerspective.ACCESSIBILITY: """
                - Test keyboard navigation
                - Verify screen reader compatibility
                - Check color contrast
                - Test focus management
                - Verify ARIA attributes and roles
            """,
            TestPerspective.PERFORMANCE: """
                - Test page load times
                - Check resource loading
                - Verify response times for interactions
                - Test under different network conditions
                - Check memory usage and leaks
            """,
            TestPerspective.USABILITY: """
                - Test interface consistency
                - Verify error message clarity
                - Check responsive design
                - Test user feedback mechanisms
                - Verify intuitive navigation
            """,
        }
        return guidance.get(perspective.name, "")

    def _create_page_prompt(
        self, page_data: dict, perspective: PerspectiveConfig
    ) -> str:
        """Create prompt for testing the entire page from a specific perspective"""
        return f"""
        Create comprehensive test cases for a web page with the following details:
        
        Page Description: {page_data.get('description', 'N/A')}
        
        Form Elements: {json.dumps(page_data.get('elements', {}), indent=2)}
        
        HTML Content: {page_data.get('html', 'N/A')}
        
        {TEST_CASES_DEFINITION_PROMPT.format(perspective=perspective.name.value)}
        
        Focus on {perspective.name.value.upper()} testing aspects including:
        1. Critical user paths
        2. Edge cases and error conditions
        3. Data validation
        4. User interactions
        5. Specific {perspective.name.value} considerations
        
        Additional considerations for {perspective.name.value} testing:
        {self._get_perspective_specific_guidance(perspective)}
        """

    def _generate_test_cases_for_page(
        self, page_data: dict
    ) -> ExpertValidatedTestsWithMetadata:
        """Generate test cases using multiple runs and expert validation"""
        all_test_results = []

        # Generate page-level test cases from different perspectives
        for perspective in TEST_PERSPECTIVES:
            prompt = self._create_page_prompt(page_data, perspective)
            result = self._generate_test_cases_with_retry(
                prompt, perspective.temperature
            )
            all_test_results.append(result)

        # Generate element-level test cases
        elements = page_data.get("elements", {})
        for element_id, element_data in elements.items():
            element_context = self._extract_element_context(
                element_id, element_data, page_data.get("html", "")
            )
            element_results = self._generate_element_test_cases(
                element_context, page_data
            )
            all_test_results.extend(element_results)

        # Have expert validate and curate all results
        return self._validate_with_expert(all_test_results, page_data)

    def _extract_element_context(
        self, element_id: str, element_data: dict | list, html_content: str
    ) -> ElementContext:
        """Extract context for a specific element including surrounding HTML"""
        # Handle case where element_data is a list
        if isinstance(element_data, list):
            return ElementContext(
                element_id=element_id,
                element_type="list",  # or some other appropriate type for list elements
                element_data={
                    "items": element_data
                },  # wrap list in a dict for consistency
                surrounding_html=html_content,
            )
        # Handle case where element_data is a dict
        elif isinstance(element_data, dict):
            return ElementContext(
                element_id=element_id,
                element_type=element_data.get("type", "unknown"),
                element_data=element_data,
                surrounding_html=html_content,
            )
        # Handle other cases
        else:
            return ElementContext(
                element_id=element_id,
                element_type="unknown",
                element_data={"value": str(element_data)},
                surrounding_html=html_content,
            )

    def _generate_element_test_cases(
        self, element_context: ElementContext, page_data: dict
    ) -> list[TestGenerationResultWithMetadata]:
        """Generate test cases for a specific element"""
        results = []

        for perspective in TEST_PERSPECTIVES:
            prompt = self._create_element_prompt(
                element_context, page_data, perspective
            )
            result = self._generate_test_cases_with_retry(
                prompt, perspective.temperature
            )
            results.append(result)

        return results

    def _create_element_prompt(
        self,
        element_context: ElementContext,
        page_data: dict,
        perspective: PerspectiveConfig,
    ) -> str:
        """Create prompt for testing a specific element"""
        return f"""
        Focus on testing the following element within the page context:

        Element Details:
        ---------------
        ID: {element_context.element_id}
        Type: {element_context.element_type}
        Properties: {json.dumps(element_context.element_data, indent=2)}

        Page Context:
        -------------
        Description: {page_data.get('description', 'N/A')}
        
        Surrounding HTML Context:
        ------------------------
        {element_context.surrounding_html}

        Generate test cases specific to this element, considering:
        1. Element's purpose and functionality
        2. Interaction with other elements
        3. Data validation and constraints
        4. Error scenarios
        5. {perspective.name.value.capitalize()} specific considerations

        {TEST_CASES_DEFINITION_PROMPT.format(perspective=perspective.name.value)}
        """

    def generate_all_test_cases(self) -> dict[str, ExpertValidatedTests]:
        """Generate test cases for all available pages"""
        test_cases = {}

        # Get all page paths from routes
        page_paths = self._extract_page_paths(self.routes)

        for page_path in page_paths:
            if (self.base_path / page_path).is_dir():
                page_data = self._load_page_data(page_path)
                if page_data:
                    test_cases[page_path] = self._generate_test_cases_for_page(
                        page_data
                    )
                    logger.info(f"Generated test cases for {page_path}")
            else:
                logger.warning(f"Directory not found: {self.base_path / page_path}")

        return test_cases


def openai_api_calculate_cost(usage, model=MODEL_NAME):
    """
    Calculate OpenAI API cost taking into account cached tokens.

    Args:
        usage: Response usage object containing token counts and details
        model: Model identifier (default: MODEL_NAME)

    Returns:
        float: Total cost in USD

    Pricing info:
    1) https://platform.openai.com/docs/pricing
    2) https://openai.com/api/pricing/
    """
    pricing_per_1m_tokens = {
        "o1": {
            "prompt": 15,
            "cached": 7.5,
            "completion": 60,
        },
        "o3-mini": {
            "prompt": 1.1,
            "cached": 0.55,
            "completion": 4.4,
        },
        "gpt-4o": {
            "prompt": 2.5,
            "cached": 1.25,
            "completion": 10,
        },
    }

    try:
        model_pricing = pricing_per_1m_tokens[model]
    except KeyError:
        raise ValueError(f"Invalid model specified: {model}")

    # Get cached tokens count
    cached_tokens = (
        usage.prompt_tokens_details.cached_tokens
        if hasattr(usage, "prompt_tokens_details")
        else 0
    )

    # Calculate non-cached prompt tokens
    non_cached_prompt_tokens = usage.prompt_tokens - cached_tokens

    # Calculate costs for each component
    prompt_cost = non_cached_prompt_tokens * model_pricing["prompt"] / 1_000_000
    cached_cost = cached_tokens * model_pricing["cached"] / 1_000_000
    completion_cost = usage.completion_tokens * model_pricing["completion"] / 1_000_000

    total_cost = prompt_cost + cached_cost + completion_cost
    total_cost = round(total_cost, 6)

    logger.info(f"\nTokens used:")
    logger.info(f"  Prompt (non-cached): {non_cached_prompt_tokens:,} tokens")
    logger.info(f"  Prompt (cached): {cached_tokens:,} tokens")
    logger.info(f"  Completion: {usage.completion_tokens:,} tokens")
    logger.info(f"  Total: {usage.total_tokens:,} tokens")
    logger.info(f"\nCosts breakdown:")
    logger.info(f"  Prompt (non-cached): ${prompt_cost:.4f}")
    logger.info(f"  Prompt (cached): ${cached_cost:.4f}")
    logger.info(f"  Completion: ${completion_cost:.4f}")
    logger.info(f"Total cost for {model}: ${total_cost:.4f}\n")

    return total_cost


def save_test_cases(test_cases: dict[str, ExpertValidatedTestsWithMetadata]):
    """
    Save generated test cases with metadata in both JSON and XLSX formats.
    Creates a separate directory for each page and saves individual files.

    Args:
        test_cases: Dictionary mapping page paths to their test cases and metadata
    """
    import json
    from datetime import datetime
    from pathlib import Path

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    def create_excel_file(test_cases_data: list, filepath: Path):
        """Create formatted Excel file from test cases data"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"

        # Define styles
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        wrap_alignment = Alignment(wrap_text=True, vertical="top")

        # Define headers
        headers = [
            "#",
            "Title",
            "Steps",
            "Expected Result",
            "Priority",
            "Category",
            "Est. Time",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Add data
        for row, test_case in enumerate(test_cases_data, 2):
            # Add test case number
            ws.cell(row=row, column=1, value=row - 1)  # Start numbering from 1

            # Add other test case data
            ws.cell(row=row, column=2, value=test_case["title"])
            ws.cell(
                row=row,
                column=3,
                value="\n".join(f"- {step}" for step in test_case["steps"]),
            )
            ws.cell(
                row=row,
                column=4,
                value="\n".join(f"- {result}" for result in test_case["result"]),
            )
            ws.cell(row=row, column=5, value=test_case["priority"])
            ws.cell(row=row, column=6, value=test_case["category"])
            ws.cell(row=row, column=7, value=test_case["estimated_time"])

        # Apply formatting
        for row in ws.iter_rows(min_row=1, max_row=len(test_cases_data) + 1):
            for cell in row:
                cell.alignment = wrap_alignment

        # Adjust column widths
        column_widths = {
            1: 8,  # #
            2: 40,  # Title
            3: 50,  # Steps
            4: 50,  # Expected Result
            5: 15,  # Priority
            6: 15,  # Category
            7: 15,  # Est. Time
        }

        for col, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        # Save the workbook
        wb.save(filepath)

    # Create output directory
    output_dir = Path(OUTPUT_PATH)
    output_dir.mkdir(exist_ok=True)

    # Prepare main test_cases.json
    output_data = {
        "generated_at": datetime.now().isoformat(),
        "model_version": MODEL_NAME,
        "pages": {
            page: {
                "test_cases": test_case.result.model_dump().get("test_cases", []),
                "validation_metadata": test_case.validation_metadata.model_dump(),
            }
            for page, test_case in test_cases.items()
        },
    }

    # Save main test_cases.json
    with open(output_dir / "test_cases.json", "w") as f:
        json.dump(output_data, f, indent=2)

    # Create individual page directories and files
    for page_path, test_case in test_cases.items():
        # Create page directory
        page_dir = output_dir / page_path
        page_dir.mkdir(exist_ok=True, parents=True)

        # Get test cases data
        page_test_cases = test_case.result.model_dump().get("test_cases", [])
        page_metadata = test_case.validation_metadata.model_dump()

        # Save page-specific JSON
        page_data = {
            "generated_at": datetime.now().isoformat(),
            "model_version": MODEL_NAME,
            "test_cases": page_test_cases,
            "metadata": page_metadata,
        }

        try:
            with open(page_dir / "test_cases.json", "w") as f:
                json.dump(page_data, f, indent=2)
        except Exception as e:
            logger.error(f"Error saving JSON for page {page_path}: {e}")

        # Create and save Excel file
        try:
            create_excel_file(page_test_cases, page_dir / "test_cases.xlsx")
        except Exception as e:
            logger.error(f"Error saving Excel file for page {page_path}: {e}")

    logger.info(f"Test cases saved successfully in {output_dir}")


def main():
    # Initialize the generator
    generator = TestCaseGenerator(INPUT_PATH)

    # Generate and validate test cases
    test_cases = generator.generate_all_test_cases()

    # Save results
    save_test_cases(test_cases)

    logger.info("Test cases generated and validated successfully!")


if __name__ == "__main__":
    main()
