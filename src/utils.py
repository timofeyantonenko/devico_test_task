from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def openai_api_calculate_cost(usage, model):
    """
    Calculate OpenAI API cost taking into account cached tokens.

    Args:
        usage: Response usage object containing token counts and details
        model: Model identifier (default: MODEL_NAME)

    Returns:
        float: Total cost in USD

    Pricing info:
    1) https://platform.openai.com/docs/pricing
    2) https://openai.com/api/pricing/
    """
    pricing_per_1m_tokens = {
        "o1": {
            "prompt": 15,
            "cached": 7.5,
            "completion": 60,
        },
        "o3-mini": {
            "prompt": 1.1,
            "cached": 0.55,
            "completion": 4.4,
        },
        "gpt-4o": {
            "prompt": 2.5,
            "cached": 1.25,
            "completion": 10,
        },
    }

    try:
        model_pricing = pricing_per_1m_tokens[model]
    except KeyError:
        raise ValueError(f"Invalid model specified: {model}")

    # Get cached tokens count
    cached_tokens = (
        usage.prompt_tokens_details.cached_tokens
        if hasattr(usage, "prompt_tokens_details")
        else 0
    )

    # Calculate non-cached prompt tokens
    non_cached_prompt_tokens = usage.prompt_tokens - cached_tokens

    # Calculate costs for each component
    prompt_cost = non_cached_prompt_tokens * model_pricing["prompt"] / 1_000_000
    cached_cost = cached_tokens * model_pricing["cached"] / 1_000_000
    completion_cost = usage.completion_tokens * model_pricing["completion"] / 1_000_000

    total_cost = prompt_cost + cached_cost + completion_cost
    total_cost = round(total_cost, 6)

    return {
        "prompt_non_cached_tokens": non_cached_prompt_tokens,
        "prompt_cached_tokens": cached_tokens,
        "completion_tokens": usage.completion_tokens,
        "total_tokens": usage.total_tokens,
        "prompt_non_cached_cost": prompt_cost,
        "prompt_cached_cost": cached_cost,
        "completion_cost": completion_cost,
        "total_cost": total_cost,
    }


def create_excel_file(test_cases_data: list, filepath: Path):
    """Create formatted Excel file from test cases data"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"  # type: ignore

    # Define styles
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    # Define headers
    headers = [
        "#",
        "Title",
        "Steps",
        "Expected Result",
        "Priority",
        "Category",
        "Est. Time",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)  # type: ignore
        cell.fill = header_fill
        cell.font = header_font

    # Add data
    for row, test_case in enumerate(test_cases_data, 2):
        # Add test case number
        # Start numbering from 1
        ws.cell(row=row, column=1, value=row - 1)  # type: ignore

        # Add other test case data
        ws.cell(row=row, column=2, value=test_case["title"])  # type: ignore
        ws.cell(  # type: ignore
            row=row,
            column=3,
            value="\n".join(f"- {step}" for step in test_case["steps"]),
        )
        ws.cell(  # type: ignore
            row=row,
            column=4,
            value="\n".join(f"- {result}" for result in test_case["result"]),
        )
        ws.cell(row=row, column=5, value=test_case["priority"])  # type: ignore
        ws.cell(row=row, column=6, value=test_case["category"])  # type: ignore
        ws.cell(row=row, column=7, value=test_case["estimated_time"])  # type: ignore

    # Apply formatting
    for row in ws.iter_rows(min_row=1, max_row=len(test_cases_data) + 1):  # type: ignore
        for cell in row:  # type: ignore
            cell.alignment = wrap_alignment

    # Adjust column widths
    column_widths = {
        1: 8,  # #
        2: 40,  # Title
        3: 50,  # Steps
        4: 50,  # Expected Result
        5: 15,  # Priority
        6: 15,  # Category
        7: 15,  # Est. Time
    }

    for col, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width  # type: ignore

    # Save the workbook
    wb.save(filepath)
